#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
公司資料查詢工具（統一編號查詢）
資料來源：
  - findbiz.nat.gov.tw    → 公司登記基本資料（全欄位，含代表人/資本額/特別股…）
  - TWSE/TPEX ISIN 清單   → 判斷上市/上櫃及股票代號
  - TWSE openapi          → 上市公司年底收盤價
  - TPEX openapi          → 上櫃公司年底收盤價
  - yfinance              → 除息 / 除權歷史資料
"""

import io
import re
import sys
import time
import urllib3
import warnings
from datetime import date as date_cls
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import requests
import yfinance as yf
from openpyxl.styles import Font
from bs4 import BeautifulSoup

from findbiz_scraper import scrape_company, search_companies_by_name

warnings.filterwarnings("ignore")
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

_REQ = {
    "headers": {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
    "verify": False,
    "timeout": 15,
}

FIND_BIZ_HOME_URL = "https://findbiz.nat.gov.tw/fts/query/QueryBar/queryInit.do"
TWSE_STOCK_DAY_URL = "https://www.twse.com.tw/rwd/zh/afterTrading/STOCK_DAY"
TPEX_STOCK_DAY_URL = "https://www.tpex.org.tw/www/zh-tw/afterTrading/tradingStock"
ESB_STOCK_DAY_URL = "https://www.tpex.org.tw/www/zh-tw/emerging/historical"
TWSE_STOCK_DAY_PAGE_URL = "https://accessibility.twse.com.tw/zh/trading/historical/stock-day.html"
TPEX_STOCK_DAY_PAGE_URL = "https://www.tpex.org.tw/zh-tw/mainboard/trading/info/stock-pricing.html"
ESB_STOCK_DAY_PAGE_URL = "https://www.tpex.org.tw/zh-tw/esb/trading/info/stock-pricing.html"
TWSE_COMPANY_PROFILE_URL = "https://openapi.twse.com.tw/v1/opendata/t187ap03_L"
TPEX_COMPANY_PROFILE_URL = "https://www.tpex.org.tw/openapi/v1/mopsfin_t187ap03_O"
ESB_COMPANY_PROFILE_URL = "https://www.tpex.org.tw/openapi/v1/mopsfin_t187ap03_R"
URL_EXPORT_COLUMNS = {"股價資料來源網址", "股價友善查詢網址", "登記資料來源網址", "列印連結"}

# ══════════════════════════════════════════════════════════════════════════════
# 1. ISIN 清單（判斷上市/上櫃）
# ══════════════════════════════════════════════════════════════════════════════

_ISIN_BY_STOCK: dict = {}   # stock_no → {stock_no, name, market}
_ISIN_BY_NAME:  list = []   # [(name, entry), ...]
_ISIN_BY_NORMALIZED_NAME: dict = {}
_OFFICIAL_BY_STOCK: dict = {}
_OFFICIAL_BY_UID: dict = {}


def _normalize_company_name(name: str) -> str:
    text = str(name or "").strip()
    text = re.sub(r"[　\s\-\(\)（）\.．,，]", "", text)
    text = re.sub(r"(股份有限公司|有限公司|公司|企業)$", "", text)
    return text


def _normalize_foreign_name(name: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(name or "").upper())


def _get_stock_name_hints(stock_no: str, market: str) -> set[str]:
    try:
        suffix = ".TW" if market == "TWSE" else ".TWO"
        info = yf.Ticker(f"{stock_no}{suffix}").info
    except Exception:
        return set()

    hints = set()
    for key in ("longName", "shortName"):
        normalized = _normalize_foreign_name(info.get(key))
        if normalized:
            hints.add(normalized)
    return hints


def load_isin() -> None:
    if _ISIN_BY_STOCK:
        return
    for mode, market in (("2", "TWSE"), ("4", "TPEX"), ("5", "ESB")):
        url = f"https://isin.twse.com.tw/isin/C_public.jsp?strMode={mode}"
        try:
            resp = requests.get(url, **{**_REQ, "timeout": 20})
            resp.encoding = "big5"
            soup = BeautifulSoup(resp.text, "html.parser")
            table = soup.find("table", {"class": "h4"})
            if not table:
                continue
            for tr in table.find_all("tr")[1:]:
                tds = tr.find_all("td")
                if not tds:
                    continue
                m = re.match(r"^(\d{3,6})\s+(.+)$", tds[0].get_text(strip=True))
                if m:
                    entry = {"stock_no": m.group(1), "name": m.group(2).strip(), "market": market}
                    _ISIN_BY_STOCK[m.group(1)] = entry
                    _ISIN_BY_NAME.append((m.group(2).strip(), entry))
                    normalized_name = _normalize_company_name(m.group(2))
                    if normalized_name:
                        _ISIN_BY_NORMALIZED_NAME.setdefault(normalized_name, []).append(entry)
        except Exception as e:
            print(f"  [警告] 無法載入 {market} ISIN：{e}")


def load_stock_profiles() -> None:
    if _OFFICIAL_BY_STOCK:
        return

    sources = [
        (
            TWSE_COMPANY_PROFILE_URL,
            "TWSE",
            {
                "stock_no": "公司代號",
                "name": "公司名稱",
                "short_name": "公司簡稱",
                "uid": "營利事業統一編號",
            },
        ),
        (
            TPEX_COMPANY_PROFILE_URL,
            "TPEX",
            {
                "stock_no": "SecuritiesCompanyCode",
                "name": "CompanyName",
                "short_name": "CompanyAbbreviation",
                "uid": "UnifiedBusinessNo.",
            },
        ),
        (
            ESB_COMPANY_PROFILE_URL,
            "ESB",
            {
                "stock_no": "SecuritiesCompanyCode",
                "name": "CompanyName",
                "short_name": "CompanyAbbreviation",
                "uid": "UnifiedBusinessNo.",
            },
        ),
    ]

    for url, market, fields in sources:
        try:
            resp = requests.get(url, **{**_REQ, "timeout": 20})
            items = resp.json()
            if not isinstance(items, list):
                continue
            for item in items:
                stock_no = str(item.get(fields["stock_no"], "")).strip()
                if not re.match(r"^\d{4,6}$", stock_no):
                    continue
                _OFFICIAL_BY_STOCK[stock_no] = {
                    "stock_no": stock_no,
                    "name": str(item.get(fields["name"], "")).strip(),
                    "short_name": str(item.get(fields["short_name"], "")).strip(),
                    "uid": str(item.get(fields["uid"], "")).strip(),
                    "market": market,
                }
                uid = str(item.get(fields["uid"], "")).strip()
                if re.fullmatch(r"\d{8}", uid):
                    _OFFICIAL_BY_UID[uid] = _OFFICIAL_BY_STOCK[stock_no]
        except Exception as e:
            print(f"  [警告] 無法載入 {market} 公司基本資料：{e}")


def _resolve_stock_entry_from_company_name(company_name: str) -> dict | None:
    normalized_name = _normalize_company_name(company_name)
    if normalized_name:
        exact_matches = [
            entry
            for entry in _ISIN_BY_NORMALIZED_NAME.get(normalized_name, [])
            if re.fullmatch(r"\d{4,6}", str(entry.get("stock_no", "")))
        ]
        if exact_matches:
            exact_matches.sort(
                key=lambda entry: (
                    len(str(entry.get("name", ""))),
                    str(entry.get("market", "")) in {"TWSE", "TPEX", "ESB"},
                ),
                reverse=True,
            )
            return exact_matches[0]

    short = normalized_name
    for name, entry in _ISIN_BY_NAME:
        sno = entry["stock_no"]
        if not re.match(r"^\d{4,6}$", sno):
            continue
        candidate_name = _normalize_company_name(name)
        if short and (short in candidate_name or candidate_name in short):
            return entry
    return None


def init_caches() -> None:
    print("  載入 ISIN 上市/上櫃清單...", end=" ", flush=True)
    load_isin()
    print(f"完成（{len(_ISIN_BY_STOCK)} 支）")
    print("  載入上市櫃公司基本資料...", end=" ", flush=True)
    load_stock_profiles()
    print(f"完成（{len(_OFFICIAL_BY_STOCK)} 筆）")


# ══════════════════════════════════════════════════════════════════════════════
# 2. 年底收盤價
# ══════════════════════════════════════════════════════════════════════════════

def _parse_price_query_date(price_date, year: int) -> date_cls:
    if isinstance(price_date, datetime):
        price_date = price_date.date()
    elif isinstance(price_date, str):
        raw = price_date.strip()
        parsed = None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                parsed = datetime.strptime(raw, fmt).date()
                break
            except ValueError:
                continue
        price_date = parsed
    elif not isinstance(price_date, date_cls):
        price_date = None

    if price_date and price_date.year == year:
        return price_date
    return date_cls(year, 12, 31)


def _get_twse_price(stock_no: str, target_date: date_cls) -> tuple:
    cursor = date_cls(target_date.year, target_date.month, 1)
    oldest = date_cls(target_date.year, 1, 1) - timedelta(days=370)
    try:
        while cursor >= oldest:
            resp = requests.get(
                TWSE_STOCK_DAY_URL,
                params={"date": cursor.strftime("%Y%m01"), "stockNo": stock_no, "response": "json"},
                **_REQ,
            )
            data = resp.json()
            if data.get("stat") == "OK" and data.get("data"):
                matched_rows = []
                for row in data["data"]:
                    roc_date = row[0]
                    parts = roc_date.split("/")
                    if len(parts) != 3:
                        continue
                    current_date = date_cls(int(parts[0]) + 1911, int(parts[1]), int(parts[2]))
                    if current_date <= target_date:
                        matched_rows.append((current_date, row))
                if matched_rows:
                    current_date, last = max(matched_rows, key=lambda item: item[0])
                    return current_date.strftime("%Y/%m/%d"), last[6]

            if cursor.month == 1:
                cursor = date_cls(cursor.year - 1, 12, 1)
            else:
                cursor = date_cls(cursor.year, cursor.month - 1, 1)
    except Exception:
        return "", ""
    return "", ""


def _extract_latest_row_from_monthly_table(rows: list, target_date: date_cls, value_index: int) -> tuple[str, str]:
    matched_rows = []
    for row in rows or []:
        if not row:
            continue
        parts = str(row[0]).split("/")
        if len(parts) != 3:
            continue
        current_date = date_cls(int(parts[0]) + 1911, int(parts[1]), int(parts[2]))
        if current_date <= target_date:
            matched_rows.append((current_date, row))
    if not matched_rows:
        return "", ""
    current_date, last_row = max(matched_rows, key=lambda item: item[0])
    value = str(last_row[value_index]).strip() if len(last_row) > value_index else ""
    return current_date.strftime("%Y/%m/%d"), value


def _get_tpex_price(stock_no: str, target_date: date_cls) -> tuple:
    cursor = date_cls(target_date.year, target_date.month, 1)
    oldest = date_cls(target_date.year, 1, 1) - timedelta(days=370)
    while cursor >= oldest:
        try:
            resp = requests.post(
                TPEX_STOCK_DAY_URL,
                data={"code": stock_no, "date": cursor.strftime("%Y/%m/01"), "response": "json"},
                **_REQ,
            )
            data = resp.json()
            tables = data.get("tables", []) or []
            rows = tables[0].get("data", []) if tables else []
            date_str, price = _extract_latest_row_from_monthly_table(rows, target_date, 6)
            if date_str and price:
                return date_str, price
        except Exception:
            time.sleep(0.3)

        if cursor.month == 1:
            cursor = date_cls(cursor.year - 1, 12, 1)
        else:
            cursor = date_cls(cursor.year, cursor.month - 1, 1)
    return "", ""


def _get_esb_price(stock_no: str, target_date: date_cls) -> tuple:
    cursor = date_cls(target_date.year, target_date.month, 1)
    oldest = date_cls(target_date.year, 1, 1) - timedelta(days=370)
    while cursor >= oldest:
        try:
            resp = requests.post(
                ESB_STOCK_DAY_URL,
                data={"code": stock_no, "date": cursor.strftime("%Y/%m/01"), "type": "Monthly", "response": "json"},
                **_REQ,
            )
            data = resp.json()
            tables = data.get("tables", []) or []
            rows = tables[0].get("data", []) if tables else []
            date_str, price = _extract_latest_row_from_monthly_table(rows, target_date, 5)
            if date_str and price and price not in {"0.00", "0", "---"}:
                return date_str, price
        except Exception:
            time.sleep(0.3)

        if cursor.month == 1:
            cursor = date_cls(cursor.year - 1, 12, 1)
        else:
            cursor = date_cls(cursor.year, cursor.month - 1, 1)
    return "", ""


def get_stock_price_on_or_before(stock_no: str, market: str, target_date: date_cls) -> tuple:
    if market == "TWSE":
        return _get_twse_price(stock_no, target_date)
    if market == "TPEX":
        return _get_tpex_price(stock_no, target_date)
    if market == "ESB":
        return _get_esb_price(stock_no, target_date)
    return "", ""


def get_stock_price_source_info(stock_no: str, market: str, year: int, date_str: str) -> tuple[str, str]:
    if not stock_no or not market:
        return "", ""
    month_param = "".join(str(date_str or "").split("/"))[:6]
    if len(month_param) != 6:
        month_param = f"{year}12"
    if market == "TWSE":
        return (
            "TWSE 個股日成交資訊（官方報表頁）",
            f"{TWSE_STOCK_DAY_URL}?date={month_param}01&stockNo={stock_no}&response=html",
        )
    if market == "TPEX":
        return (
            "TPEX 上櫃個股歷史行情（官方查詢頁）",
            f"https://www.tpex.org.tw/zh-tw/mainboard/trading/info/stock-pricing.html?code={stock_no}&date={month_param}",
        )
    if market == "ESB":
        return (
            "TPEX 興櫃個股歷史行情（成交均價，官方查詢頁）",
            f"{ESB_STOCK_DAY_PAGE_URL}?code={stock_no}&date={month_param}",
        )
    return "", ""


def get_stock_query_page_info(market: str) -> tuple[str, str]:
    if market == "TWSE":
        return ("TWSE 友善查詢頁", TWSE_STOCK_DAY_PAGE_URL)
    if market == "TPEX":
        return ("TPEX 上櫃資料獲取來源", TPEX_STOCK_DAY_PAGE_URL)
    if market == "ESB":
        return ("TPEX 興櫃資料獲取來源", ESB_STOCK_DAY_PAGE_URL)
    return "", ""


# ══════════════════════════════════════════════════════════════════════════════
# 3. 除權息資訊（yfinance）
# ══════════════════════════════════════════════════════════════════════════════

def get_dividend_window_label(year: int, years_back: int = 2) -> str:
    start_year = year - years_back + 1
    return f"{start_year}-{year}"


def get_dividends(stock_no: str, market: str, year: int, years_back: int = 2) -> list:
    try:
        suffix = ".TW" if market == "TWSE" else ".TWO"
        tk = yf.Ticker(f"{stock_no}{suffix}")
        records = {}
        target_years = set(range(year - years_back + 1, year + 1))

        divs = tk.dividends
        if not divs.empty:
            for ts, amount in divs.items():
                if ts.year not in target_years:
                    continue
                d = ts.strftime("%Y-%m-%d")
                records.setdefault(d, {"日期": d, "類別": "除息",
                                       "現金股利(元)": "", "股票股利(元)": ""})
                records[d]["現金股利(元)"] = round(float(amount), 4)

        splits = tk.splits
        if not splits.empty:
            for ts, ratio in splits.items():
                if ts.year not in target_years or ratio == 1.0:
                    continue
                d = ts.strftime("%Y-%m-%d")
                stock_div = round((float(ratio) - 1) * 10, 4)
                if d in records:
                    records[d]["類別"] = "除權息"
                    records[d]["股票股利(元)"] = stock_div
                else:
                    records[d] = {"日期": d, "類別": "除權",
                                  "現金股利(元)": "", "股票股利(元)": stock_div}

        return sorted(records.values(), key=lambda x: x["日期"])
    except Exception:
        return []


# ══════════════════════════════════════════════════════════════════════════════
# 4. 單一公司完整查詢（統一編號）
# ══════════════════════════════════════════════════════════════════════════════

RESULT_COLUMNS = [
    # ── 公司登記基本資料（findbiz）
    "統一編號", "登記現況", "股權狀況",
    "公司名稱", "章程所訂外文公司名稱",
    "資本總額(元)", "實收資本額(元)",
    "每股金額(元)", "已發行股份總數(股)",
    "代表人姓名", "公司所在地", "登記機關",
    "核准設立日期", "最後核准變更日期",
    "複數表決權特別股", "對於特定事項具否決權特別股", "特別股董監選任限制",
    "董監事任期", "董監事資料",
    "所營事業",
    # ── 股市資訊
    "股票代號", "市場別",
    "股價查詢日期",
    "實際收盤日期", "收盤價(元)",
    "股價資料來源說明", "股價資料來源網址",
    "股價友善查詢說明", "股價友善查詢網址",
    # ── 除權息
    "除權息查詢區間",
    "除權息筆數", "除權息明細",
    # ── 連結
    "登記資料來源網址", "列印連結",
    "備註",
]


def get_trading_days(year: int) -> list[date_cls]:
    days: set[date_cls] = set()
    for month in range(1, 13):
        try:
            resp = requests.get(
                TWSE_STOCK_DAY_URL,
                params={"date": f"{year}{month:02d}01", "stockNo": "2330", "response": "json"},
                **_REQ,
            )
            data = resp.json()
            for row in data.get("data", []) or []:
                parts = str(row[0]).split("/")
                if len(parts) != 3:
                    continue
                current_date = date_cls(int(parts[0]) + 1911, int(parts[1]), int(parts[2]))
                if current_date.year == year:
                    days.add(current_date)
        except Exception:
            continue
    return sorted(days, reverse=True)


def _score_stock_candidate(candidate_name: str, target_name: str) -> int:
    candidate_text = str(candidate_name or "").strip()
    target_text = str(target_name or "").strip()
    candidate_norm = _normalize_company_name(candidate_text)
    target_norm = _normalize_company_name(target_text)
    score = 0
    if candidate_text == target_text:
        score += 120
    if candidate_norm == target_norm:
        score += 80
    if target_norm and target_norm in candidate_norm:
        score += 40
    if candidate_norm and candidate_norm in target_norm:
        score += 25
    if "股份有限公司" in candidate_text:
        score += 15
    return score


def query_by_name(company_name: str, year: int, price_date=None) -> dict:
    result = {col: "" for col in RESULT_COLUMNS}
    result["年底收盤日期"] = ""
    result["年底收盤價(元)"] = ""
    name = str(company_name or "").strip()
    result["公司名稱"] = name
    result["股價查詢日期"] = _parse_price_query_date(price_date, year).strftime("%Y/%m/%d")

    if not name:
        result["備註"] = "請輸入公司名稱"
        return result

    candidates = search_companies_by_name(name)
    if not candidates:
        result["備註"] = "查無符合公司名稱，請確認名稱或改用統一編號/股票代號查詢"
        return result

    ordered = sorted(
        candidates,
        key=lambda item: (_score_stock_candidate(item.get("name", ""), name), item.get("ban", "")),
        reverse=True,
    )
    chosen = ordered[0]
    resolved = query_by_uid(chosen["ban"], year, price_date=price_date)
    if len(ordered) > 1 and resolved.get("公司名稱"):
        note = f"公司名稱搜尋共 {len(ordered)} 筆，已自動採用最相符結果（統編 {chosen['ban']}）"
        resolved["備註"] = f"{resolved['備註']}；{note}" if resolved.get("備註") else note
    return resolved


def _resolve_uid_from_stock_no(stock_no: str) -> tuple[list[dict], dict | None, str]:
    load_isin()
    load_stock_profiles()
    normalized_stock_no = str(stock_no or "").strip()
    official_entry = _OFFICIAL_BY_STOCK.get(normalized_stock_no)
    if official_entry and official_entry.get("uid"):
        return [{"ban": official_entry["uid"], "name": official_entry.get("name") or official_entry.get("short_name") or normalized_stock_no}], official_entry, ""

    entry = official_entry or _ISIN_BY_STOCK.get(normalized_stock_no)
    if not entry:
        return [], None, "查無此股票代號，請確認是否正確"

    target_name = entry.get("name", "") or entry.get("short_name", "")
    candidates = search_companies_by_name(target_name)
    if not candidates:
        return [], entry, f"已找到股票代號 {normalized_stock_no}，但暫時無法對應公司統編"

    ordered = sorted(
        candidates,
        key=lambda item: (_score_stock_candidate(item.get("name", ""), target_name), item.get("ban", "")),
        reverse=True,
    )
    return ordered, entry, ""


def query_by_uid(unified_id: str, year: int, price_date=None) -> dict:
    load_isin()
    load_stock_profiles()
    result = {col: "" for col in RESULT_COLUMNS}
    result["年底收盤日期"] = ""
    result["年底收盤價(元)"] = ""
    uid = unified_id.strip()
    result["統一編號"] = uid
    target_date = _parse_price_query_date(price_date, year)
    result["股價查詢日期"] = target_date.strftime("%Y/%m/%d")

    # Step 1: findbiz 公司登記資料
    fb = scrape_company(uid)
    if fb.get("_error"):
        result["備註"] = fb["_error"]
        return result
    if not fb.get("公司名稱"):
        result["備註"] = "查無此統一編號，請確認是否正確"
        return result

    # 填入 findbiz 欄位
    for col in ["統一編號", "登記現況", "股權狀況",
                "公司名稱", "章程所訂外文公司名稱",
                "資本總額(元)", "實收資本額(元)",
                "每股金額(元)", "已發行股份總數(股)",
                "代表人姓名", "公司所在地", "登記機關",
                "核准設立日期", "最後核准變更日期",
                "複數表決權特別股", "對於特定事項具否決權特別股", "特別股董監選任限制",
                "董監事任期", "董監事資料",
                "所營事業"]:
        result[col] = fb.get(col, "")

    stable_findbiz_url = fb.get("_share_url", "") or f"{FIND_BIZ_HOME_URL}?banNo={uid}"
    result["登記資料來源網址"] = stable_findbiz_url
    result["列印連結"]         = stable_findbiz_url
    result["_detail_html"]    = fb.get("_detail_html", "")
    result["_snapshot_at"]    = fb.get("_snapshot_at", "")

    # Step 2: 尋找股票代號（ISIN）
    # 只比對 4~6 碼的普通股（過濾權證 6 碼含字母的代號）
    co_name = fb["公司名稱"]
    stock_no = None
    market = None

    official_entry = _OFFICIAL_BY_UID.get(uid)
    if official_entry:
        stock_no = str(official_entry.get("stock_no", "")).strip()
        market = str(official_entry.get("market", "")).strip()
    else:
        matched_entry = _resolve_stock_entry_from_company_name(co_name)
        if matched_entry:
            stock_no = str(matched_entry.get("stock_no", "")).strip()
            market = str(matched_entry.get("market", "")).strip()

    # Step 3: 年底收盤價
    if stock_no and market:
        result["股票代號"] = stock_no
        result["市場別"]   = (
            "上市(TWSE)" if market == "TWSE"
            else "上櫃(TPEX)" if market == "TPEX"
            else "興櫃(ESB)" if market == "ESB"
            else market
        )
        date_str, price = get_stock_price_on_or_before(stock_no, market, target_date)
        result["實際收盤日期"] = date_str
        result["收盤價(元)"] = price
        result["年底收盤日期"]   = date_str
        result["年底收盤價(元)"] = price
        price_source_label, price_source_url = get_stock_price_source_info(stock_no, market, year, date_str)
        query_page_label, query_page_url = get_stock_query_page_info(market)
        result["股價資料來源說明"] = price_source_label
        result["股價資料來源網址"] = price_source_url
        result["股價友善查詢說明"] = query_page_label
        result["股價友善查詢網址"] = query_page_url
        if market == "ESB":
            esb_note = "興櫃市場官方歷史行情提供的是成交均價，已據此呈現股價資訊"
            result["備註"] = f"{result['備註']}；{esb_note}" if result.get("備註") else esb_note

        # Step 4: 除權息
        result["除權息查詢區間"] = get_dividend_window_label(year)
        divs = get_dividends(stock_no, market, year, years_back=2)
        result["除權息筆數"] = str(len(divs)) if divs else "0"
        result["除權息明細"] = divs
    else:
        result["市場別"] = "未上市/未上櫃/未登錄興櫃"

    return result


def query_by_stock_no(stock_no: str, year: int, price_date=None) -> dict:
    result = {col: "" for col in RESULT_COLUMNS}
    result["年底收盤日期"] = ""
    result["年底收盤價(元)"] = ""
    normalized_stock_no = str(stock_no or "").strip()
    result["股票代號"] = normalized_stock_no
    result["股價查詢日期"] = _parse_price_query_date(price_date, year).strftime("%Y/%m/%d")

    if not re.match(r"^\d{4,6}$", normalized_stock_no):
        result["備註"] = "請輸入正確的 4 至 6 碼股票代號"
        return result

    candidates, entry, note = _resolve_uid_from_stock_no(normalized_stock_no)
    if not candidates:
        if entry and entry.get("market"):
            market = entry.get("market", "")
            target_date = _parse_price_query_date(price_date, year)
            date_str, price = get_stock_price_on_or_before(normalized_stock_no, market, target_date)
            label, url = get_stock_price_source_info(normalized_stock_no, market, year, date_str)
            query_page_label, query_page_url = get_stock_query_page_info(market)
            result["公司名稱"] = entry.get("name") or entry.get("short_name") or ""
            result["市場別"] = (
                "上市(TWSE)" if market == "TWSE"
                else "上櫃(TPEX)" if market == "TPEX"
                else "興櫃(ESB)" if market == "ESB"
                else market
            )
            result["實際收盤日期"] = date_str
            result["收盤價(元)"] = price
            result["年底收盤日期"] = date_str
            result["年底收盤價(元)"] = price
            result["股價資料來源說明"] = label
            result["股價資料來源網址"] = url
            result["股價友善查詢說明"] = query_page_label
            result["股價友善查詢網址"] = query_page_url
            divs = get_dividends(normalized_stock_no, market, year, years_back=2)
            result["除權息查詢區間"] = get_dividend_window_label(year)
            result["除權息筆數"] = str(len(divs)) if divs else "0"
            result["除權息明細"] = divs
            fallback_note = note or f"已依股票代號 {normalized_stock_no} 載入市場與股價資料，但公司登記資料暫時無法唯一對應統編"
            if market == "ESB":
                esb_note = "興櫃市場官方歷史行情提供的是成交均價，已據此呈現股價資訊"
                fallback_note = f"{fallback_note}；{esb_note}"
            result["備註"] = fallback_note
        else:
            result["市場別"] = ""
            result["備註"] = note or "查無此股票代號對應公司"
        return result

    fallback_resolved = None
    name_hints = _get_stock_name_hints(normalized_stock_no, entry.get("market", "") if entry else "")
    for candidate in candidates[:8]:
        raw_fb = scrape_company(candidate["ban"])
        foreign_norm = _normalize_foreign_name(raw_fb.get("章程所訂外文公司名稱"))
        if name_hints and foreign_norm:
            if any(
                foreign_norm == hint or foreign_norm in hint or hint in foreign_norm
                for hint in name_hints
            ):
                return query_by_uid(candidate["ban"], year, price_date=price_date)

        resolved = query_by_uid(candidate["ban"], year, price_date=price_date)
        if not resolved.get("公司名稱"):
            continue
        if fallback_resolved is None:
            fallback_resolved = resolved
        if resolved.get("股票代號") == normalized_stock_no:
            return resolved

    if fallback_resolved and entry:
        market = entry.get("market", "")
        target_date = _parse_price_query_date(price_date, year)
        date_str, price = get_stock_price_on_or_before(normalized_stock_no, market, target_date)
        fallback_resolved["股票代號"] = normalized_stock_no
        fallback_resolved["市場別"] = (
            "上市(TWSE)" if market == "TWSE"
            else "上櫃(TPEX)" if market == "TPEX"
            else "興櫃(ESB)" if market == "ESB"
            else fallback_resolved.get("市場別", "")
        )
        fallback_resolved["股價查詢日期"] = target_date.strftime("%Y/%m/%d")
        fallback_resolved["實際收盤日期"] = date_str
        fallback_resolved["收盤價(元)"] = price
        fallback_resolved["年底收盤日期"] = date_str
        fallback_resolved["年底收盤價(元)"] = price
        label, url = get_stock_price_source_info(normalized_stock_no, market, year, date_str)
        query_page_label, query_page_url = get_stock_query_page_info(market)
        fallback_resolved["股價資料來源說明"] = label
        fallback_resolved["股價資料來源網址"] = url
        fallback_resolved["股價友善查詢說明"] = query_page_label
        fallback_resolved["股價友善查詢網址"] = query_page_url
        if market == "ESB":
            esb_note = "興櫃市場官方歷史行情提供的是成交均價，已據此呈現股價資訊"
            fallback_resolved["備註"] = f"{fallback_resolved['備註']}；{esb_note}" if fallback_resolved.get("備註") else esb_note
        divs = get_dividends(normalized_stock_no, market, year, years_back=2)
        fallback_resolved["除權息查詢區間"] = get_dividend_window_label(year)
        fallback_resolved["除權息筆數"] = str(len(divs)) if divs else "0"
        fallback_resolved["除權息明細"] = divs
        fallback_resolved["備註"] = (
            fallback_resolved.get("備註")
            or f"已依股票代號 {normalized_stock_no} 補齊股價與除權息資料，但公司名稱對應結果可能需要人工複核"
        )
        return fallback_resolved

    result["市場別"] = (
        "上市(TWSE)" if entry and entry.get("market") == "TWSE"
        else "上櫃(TPEX)" if entry and entry.get("market") == "TPEX"
        else "興櫃(ESB)" if entry and entry.get("market") == "ESB"
        else ""
    )
    result["備註"] = f"已找到股票代號 {normalized_stock_no}，但暫時無法唯一對應公司統編，建議改用統編或公司名稱查詢"
    return result


def _header_matches_uid(text: str) -> bool:
    normalized = re.sub(r"[\s_（）()\-]", "", str(text or "").lower())
    return (
        "統一編號" in normalized
        or "統編" in normalized
        or ("統一" in normalized and "編" in normalized)
        or normalized in {"uid", "ban", "businessno", "unifiedbusinessno"}
    )


def _header_matches_stock(text: str) -> bool:
    normalized = re.sub(r"[\s_（）()\-]", "", str(text or "").lower())
    return (
        "股票代號" in normalized
        or "股號" in normalized
        or ("股票" in normalized and "代號" in normalized)
        or normalized in {"stock", "stockno", "ticker", "symbol"}
    )


def _header_matches_name(text: str) -> bool:
    normalized = re.sub(r"[\s_（）()\-]", "", str(text or "").lower())
    return (
        "公司名稱" in normalized
        or ("公司" in normalized and "名稱" in normalized)
        or normalized in {"name", "company", "companyname"}
    )


def _normalize_cell_value(value) -> str:
    text = str(value or "").strip()
    return "" if not text or text.lower() == "nan" else text


def _first_nonempty(row: pd.Series, columns: list[str]) -> str:
    for column in columns:
        value = _normalize_cell_value(row.get(column, ""))
        if value:
            return value
    return ""


def _infer_query_type(value: str) -> str:
    if re.fullmatch(r"\d{8}", value or ""):
        return "uid"
    if re.fullmatch(r"\d{4,6}", value or ""):
        return "stock"
    return "name" if value else ""


def extract_batch_requests(df: pd.DataFrame) -> tuple[list[dict], list[str]]:
    columns = [str(col) for col in df.columns]
    df = df.copy()
    df.columns = columns
    uid_cols = [col for col in columns if _header_matches_uid(col)]
    stock_cols = [col for col in columns if _header_matches_stock(col)]
    name_cols = [col for col in columns if _header_matches_name(col)]
    fallback_cols = [columns[0]] if columns else []

    requests_list: list[dict] = []
    used_columns: list[str] = []

    for index, row in df.iterrows():
        uid_value = _first_nonempty(row, uid_cols)
        stock_value = _first_nonempty(row, stock_cols)
        name_value = _first_nonempty(row, name_cols)

        query_type = ""
        query_value = ""
        source_column = ""

        if uid_value:
            query_type, query_value = "uid", uid_value
            source_column = uid_cols[0]
        elif stock_value:
            query_type, query_value = "stock", stock_value
            source_column = stock_cols[0]
        elif name_value:
            query_type, query_value = "name", name_value
            source_column = name_cols[0]
        elif fallback_cols:
            query_value = _first_nonempty(row, fallback_cols)
            query_type = _infer_query_type(query_value)
            source_column = fallback_cols[0] if query_value else ""

        if not query_type or not query_value:
            continue

        requests_list.append(
            {
                "query_type": query_type,
                "query_value": query_value,
                "source_column": source_column,
                "row_number": int(index) + 2,
            }
        )
        if source_column and source_column not in used_columns:
            used_columns.append(source_column)

    return requests_list, used_columns


def run_batch_request(request_item: dict, year: int, price_date=None) -> dict:
    query_type = request_item.get("query_type")
    query_value = str(request_item.get("query_value", "")).strip()
    if query_type == "uid":
        return query_by_uid(query_value, year, price_date=price_date)
    if query_type == "stock":
        return query_by_stock_no(query_value, year, price_date=price_date)
    if query_type == "name":
        return query_by_name(query_value, year, price_date=price_date)

    result = {col: "" for col in RESULT_COLUMNS}
    result["備註"] = "不支援的批次查詢類型"
    return result


# ══════════════════════════════════════════════════════════════════════════════
# 5. 批量查詢
# ══════════════════════════════════════════════════════════════════════════════

def batch_query(input_file: str, year: int, price_date=None) -> list:
    path = Path(input_file)
    if not path.exists():
        print(f"[錯誤] 找不到檔案：{input_file}")
        sys.exit(1)

    df = pd.read_excel(input_file, dtype=str) if path.suffix.lower() in (".xlsx", ".xls") \
        else pd.read_csv(input_file, dtype=str)

    requests_list, used_columns = extract_batch_requests(df)
    columns_text = "、".join(used_columns) if used_columns else (str(df.columns[0]) if len(df.columns) else "（無）")
    print(f"  讀取欄位「{columns_text}」，共 {len(requests_list)} 筆有效查詢\n")

    init_caches()
    results = []
    for i, request_item in enumerate(requests_list, start=1):
        val = request_item["query_value"]
        label_map = {"uid": "統編", "stock": "股號", "name": "名稱"}
        query_label = label_map.get(request_item["query_type"], "查詢")
        print(f"  [{i:>3}/{len(requests_list)}] {query_label}:{val}", end=" ... ", flush=True)
        res = run_batch_request(request_item, year, price_date=price_date)
        print(res.get("公司名稱") or res.get("備註") or "（查無）")
        results.append(res)
        time.sleep(1.0)   # findbiz 有 rate limit
    return results


# ══════════════════════════════════════════════════════════════════════════════
# 6. 匯出：Excel / CSV
# ══════════════════════════════════════════════════════════════════════════════

def _flatten_result(res: dict) -> dict:
    flat = dict(res)
    biz = flat.get("所營事業", "")
    flat["所營事業"] = biz if isinstance(biz, str) else "\n".join(biz)
    directors = flat.get("董監事資料", [])
    if isinstance(directors, list) and directors:
        flat["董監事資料"] = "\n".join(
            f"{d.get('序號','')} {d.get('職稱','')} {d.get('姓名','')}"
            f" 代表法人:{d.get('所代表法人','—') or '—'}"
            f" 持股:{d.get('持有股份數(股)','—') or '—'}"
            for d in directors
        )
    else:
        flat["董監事資料"] = ""
    divs = flat.get("除權息明細", [])
    if isinstance(divs, list) and divs:
        flat["除權息明細"] = "\n".join(
            f"{d.get('日期','')} [{d.get('類別','')}]"
            f" 現金:{d.get('現金股利(元)','—')}元"
            f" 股票:{d.get('股票股利(元)','—')}元"
            for d in divs
        )
    else:
        flat["除權息明細"] = ""
    return flat


def results_to_df(results: list) -> pd.DataFrame:
    flat = [_flatten_result(r) for r in results]
    return pd.DataFrame(flat, columns=RESULT_COLUMNS)


def _format_excel_worksheet(ws) -> None:
    hyperlink_font = Font(color="0563C1", underline="single")
    header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        has_url = any(isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")) for cell in col_cells[1:])
        width_limit = 80 if has_url else 50
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, width_limit)

        for cell in col_cells[1:]:
            if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                target = cell.value
                header = ws.cell(row=1, column=cell.column).value
                if header == "股價資料來源網址":
                    label_col = header_map.get("股價資料來源說明")
                    display = ws.cell(row=cell.row, column=label_col).value if label_col else None
                    cell.value = display or "查看股價來源"
                elif header == "股價友善查詢網址":
                    label_col = header_map.get("股價友善查詢說明")
                    display = ws.cell(row=cell.row, column=label_col).value if label_col else None
                    cell.value = display or "查看股價友善查詢頁"
                elif header == "登記資料來源網址":
                    cell.value = "查看登記資料"
                elif header == "列印連結":
                    cell.value = "查看來源頁面"
                cell.hyperlink = target
                cell.font = hyperlink_font
    ws.freeze_panes = "A2"


def save_excel(results: list, output_file: str) -> None:
    df = results_to_df(results)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="查詢結果")
        ws = writer.sheets["查詢結果"]
        _format_excel_worksheet(ws)
    print(f"\n結果已儲存至：{output_file}")


def to_excel_bytes(results: list) -> bytes:
    buf = io.BytesIO()
    df = results_to_df(results)
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="查詢結果")
        ws = writer.sheets["查詢結果"]
        _format_excel_worksheet(ws)
    return buf.getvalue()


def to_csv_bytes(results: list) -> bytes:
    df = results_to_df(results)
    for col in URL_EXPORT_COLUMNS:
        if col in df.columns:
            def _to_hyperlink_formula(row):
                value = row[col]
                if not (isinstance(value, str) and value.startswith(("http://", "https://"))):
                    return value
                if col == "股價資料來源網址":
                    display_text = row.get("股價資料來源說明") or "查看股價來源"
                elif col == "股價友善查詢網址":
                    display_text = row.get("股價友善查詢說明") or "查看股價友善查詢頁"
                elif col == "登記資料來源網址":
                    display_text = "查看登記資料"
                else:
                    display_text = "查看來源頁面"
                target = str(value).replace('"', '""')
                display = str(display_text).replace('"', '""')
                return f'=HYPERLINK("{target}", "{display}")'

            df[col] = df.apply(_to_hyperlink_formula, axis=1)
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


# ══════════════════════════════════════════════════════════════════════════════
# 7. CLI
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(description="公司資料查詢工具（統一編號）")
    grp = parser.add_mutually_exclusive_group(required=True)
    grp.add_argument("--id",    "-i", dest="unified_id", metavar="統編")
    grp.add_argument("--stock", "-s", dest="stock_no", metavar="股號")
    grp.add_argument("--batch", "-b", metavar="檔案")
    parser.add_argument("--year",   "-y", type=int, default=datetime.now().year - 1)
    parser.add_argument("--price-date", dest="price_date", metavar="YYYY-MM-DD")
    parser.add_argument("--output", "-o", metavar="輸出檔")
    args = parser.parse_args()

    year = args.year
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = args.output or f"查詢結果_{ts}.xlsx"

    print(f"\n{'='*50}\n  公司資料查詢工具  ／  年份：{year}\n{'='*50}\n")

    if args.batch:
        results = batch_query(args.batch, year, price_date=args.price_date)
    elif args.stock_no:
        init_caches()
        res = query_by_stock_no(args.stock_no, year, price_date=args.price_date)
        results = [res]
    else:
        init_caches()
        res = query_by_uid(args.unified_id, year, price_date=args.price_date)
        results = [res]
        flat = _flatten_result(res)
        print("\n查詢結果：")
        for k, v in flat.items():
            if v:
                print(f"  {k:<22}: {str(v)[:60]}")

    save_excel(results, output_file)


if __name__ == "__main__":
    main()
