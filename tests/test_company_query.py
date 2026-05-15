import unittest
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook
import pandas as pd

import company_query
import findbiz_scraper


class _FakeResponse:
    def __init__(self, text):
        self.text = text
        self.encoding = "utf-8"


class CompanyQueryTests(unittest.TestCase):
    def setUp(self):
        company_query._ISIN_BY_STOCK.clear()
        company_query._ISIN_BY_NAME.clear()
        company_query._ISIN_BY_NORMALIZED_NAME.clear()
        company_query._OFFICIAL_BY_STOCK.clear()
        company_query._OFFICIAL_BY_UID.clear()

    def test_load_isin_keeps_etf_metadata(self):
        html = """
        <table class="h4">
          <tr><td>有價證券代號及名稱</td><td>國際證券辨識號碼(ISIN Code)</td><td>上市日</td><td>市場別</td><td>產業別</td><td>CFICode</td><td>備註</td></tr>
          <tr><td>ETF</td></tr>
          <tr><td>00679B　元大美債20年</td><td>TW00000679B0</td><td>2017/01/17</td><td>上櫃</td><td></td><td>CEOIBU</td><td></td></tr>
        </table>
        """

        def fake_get(url, **kwargs):
            if "strMode=4" in url:
                return _FakeResponse(html)
            return _FakeResponse("<html></html>")

        with patch("company_query.requests.get", side_effect=fake_get):
            company_query.load_isin()

        entry = company_query._ISIN_BY_STOCK["00679B"]
        self.assertEqual(entry["security_type"], "ETF")
        self.assertEqual(entry["issue_country"], "台灣")
        self.assertEqual(entry["isin_code"], "TW00000679B0")

        result = {}
        company_query._apply_security_metadata(result, entry)
        self.assertEqual(result["發行地查詢說明"], "發行地：台灣（查看鉅亨網個股頁）")
        self.assertEqual(result["發行地查詢網址"], "https://www.cnyes.com/twstock/00679B")
        self.assertEqual(result["ISIN資料來源說明"], "查看 TWSE ISIN 公開資料")
        self.assertEqual(
            result["ISIN資料來源網址"],
            "https://isin.twse.com.tw/isin/C_public.jsp?strMode=4",
        )

    def test_cnyes_stock_page_is_friendly_stock_link(self):
        label, url = company_query.get_stock_query_page_info("TWSE", "2330")

        self.assertEqual(label, "鉅亨網個股頁")
        self.assertEqual(url, "https://www.cnyes.com/twstock/2330")

    def test_batch_infers_alphanumeric_etf_code_as_stock(self):
        df = pd.DataFrame({"query": ["00679B"]})

        requests_list, _ = company_query.extract_batch_requests(df)

        self.assertEqual(requests_list[0]["query_type"], "stock")
        self.assertEqual(requests_list[0]["query_value"], "00679B")

    def test_excel_export_keeps_complete_source_links(self):
        result = {column: "" for column in company_query.RESULT_COLUMNS}
        result.update(
            {
                "公司名稱": "測試公司",
                "股票代號": "2330",
                "股價資料來源說明": "TWSE 個股日成交資訊（官方報表）",
                "股價資料來源網址": "https://example.test/price",
                "ISIN資料來源說明": "查看 TWSE ISIN 公開資料",
                "ISIN資料來源網址": "https://example.test/isin",
                "股價友善查詢說明": "TWSE 友善查詢頁",
                "股價友善查詢網址": "https://example.test/friendly",
                "公司登記資料說明": "查看 findbiz 官方頁面",
                "登記資料來源網址": "https://example.test/findbiz",
                "除權息資料來源說明": "查看 Yahoo 股利頁；查看 MOPS 查詢頁",
                "Yahoo股利頁網址": "https://example.test/yahoo",
                "MOPS查詢頁網址": "https://example.test/mops",
            }
        )

        workbook = load_workbook(BytesIO(company_query.to_excel_bytes([result])))
        sheet = workbook["查詢結果"]
        headers = [cell.value for cell in sheet[1]]
        row = {header: sheet.cell(row=2, column=index + 1) for index, header in enumerate(headers)}

        self.assertEqual(row["登記資料來源網址"].value, "查看 findbiz 官方頁面")
        self.assertEqual(row["ISIN資料來源網址"].value, "查看 TWSE ISIN 公開資料")
        self.assertEqual(row["股價資料來源網址"].value, "TWSE 個股日成交資訊（官方報表）")
        self.assertEqual(row["股價友善查詢網址"].value, "TWSE 友善查詢頁")
        self.assertEqual(row["Yahoo股利頁網址"].value, "查看 Yahoo 股利頁")
        self.assertEqual(row["MOPS查詢頁網址"].value, "查看 MOPS 查詢頁")
        self.assertEqual(row["Yahoo股利頁網址"].hyperlink.target, "https://example.test/yahoo")
        self.assertEqual(row["MOPS查詢頁網址"].hyperlink.target, "https://example.test/mops")
        self.assertEqual(row["ISIN資料來源網址"].hyperlink.target, "https://example.test/isin")

    def test_stock_lookup_uses_official_profile_when_findbiz_is_blocked(self):
        profile_entry = {
            "stock_no": "2834",
            "name": "臺灣中小企業銀行股份有限公司",
            "short_name": "臺企銀",
            "uid": "03793407",
            "market": "TWSE",
            "raw": {
                "公司代號": "2834",
                "公司名稱": "臺灣中小企業銀行股份有限公司",
                "公司簡稱": "臺企銀",
                "營利事業統一編號": "03793407",
                "董事長": "李嘉祥",
                "住址": "台北市塔城街三十號",
                "成立日期": "19500923",
                "實收資本額": "97180618490",
                "普通股每股面額": "新台幣 10.0000元",
                "已發行普通股數或TDR原股發行股數": "9718061849",
            },
        }
        isin_entry = {
            **profile_entry,
            "isin_code": "TW0002834009",
            "issue_country": "台灣",
            "security_type": "股票",
        }
        company_query._OFFICIAL_BY_STOCK["2834"] = profile_entry
        company_query._OFFICIAL_BY_UID["03793407"] = profile_entry
        company_query._ISIN_BY_STOCK["2834"] = isin_entry

        with patch("company_query.scrape_company", return_value={"統一編號": "03793407", "公司名稱": "", "_error": "403 Client Error"}), \
             patch("company_query.get_stock_price_on_or_before", return_value=("2026/05/13", "18.80")), \
             patch("company_query.get_dividends", return_value=[]):
            result = company_query.query_by_stock_no("2834", 2026, price_date="2026/05/13")

        self.assertEqual(result["統一編號"], "03793407")
        self.assertEqual(result["公司名稱"], "臺灣中小企業銀行股份有限公司")
        self.assertEqual(result["代表人姓名"], "李嘉祥")
        self.assertEqual(result["公司所在地"], "台北市塔城街三十號")
        self.assertEqual(result["核准設立日期"], "1950/09/23")
        self.assertEqual(result["實收資本額(元)"], "97180618490")
        self.assertEqual(result["股票代號"], "2834")
        self.assertEqual(result["市場別"], "上市(TWSE)")
        self.assertEqual(result["實際收盤日期"], "2026/05/13")
        self.assertEqual(result["股權狀況"], "開放資料未提供")
        self.assertEqual(result["公司登記資料說明"], "查看 findbiz 官方頁面")
        self.assertEqual(
            result["登記資料來源網址"],
            "https://findbiz.nat.gov.tw/fts/query/QueryBar/queryInit.do?banNo=03793407",
        )
        self.assertNotIn("openapi", result["登記資料來源網址"])
        self.assertIn("findbiz", result["備註"])

    def test_scrape_company_falls_back_to_gcis_open_data(self):
        def fake_get(url, **kwargs):
            if "findbiz.nat.gov.tw" in url:
                raise RuntimeError("403 Client Error")

            class _JsonResponse:
                def __init__(self, payload):
                    self._payload = payload

                def raise_for_status(self):
                    return None

                def json(self):
                    return self._payload

            if "5F64D864" in url:
                return _JsonResponse([
                    {
                        "Business_Accounting_NO": "34051920",
                        "Company_Status_Desc": "核准設立",
                        "Company_Name": "台達電子工業股份有限公司",
                        "Capital_Stock_Amount": 60000000000,
                        "Paid_In_Capital_Amount": 24567776290,
                        "Responsible_Name": "鄭平",
                        "Company_Location": "桃園市龜山區山頂里興邦路31之1號",
                        "Register_Organization_Desc": "經濟部商業發展署",
                        "Company_Setup_Date": "0640820",
                        "Change_Of_Approval_Data": "1150410",
                    }
                ])
            if "236EE382" in url:
                return _JsonResponse([
                    {
                        "Cmp_Business": [
                            {"Business_Item": "CC01080", "Business_Item_Desc": "電子零組件製造業"},
                        ]
                    }
                ])
            return _JsonResponse([])

        with patch("findbiz_scraper.requests.get", side_effect=fake_get):
            result = findbiz_scraper.scrape_company("34051920")

        self.assertEqual(result["統一編號"], "34051920")
        self.assertEqual(result["公司名稱"], "台達電子工業股份有限公司")
        self.assertEqual(result["代表人姓名"], "鄭平")
        self.assertEqual(result["核准設立日期"], "1975/08/20")
        self.assertIn("CC01080", result["所營事業"])
        self.assertEqual(
            result["_share_url"],
            "https://findbiz.nat.gov.tw/fts/query/QueryBar/queryInit.do?banNo=34051920",
        )

    def test_query_by_uid_merges_gcis_and_official_profile_without_raw_json_link(self):
        profile_entry = {
            "stock_no": "3016",
            "name": "嘉晶電子股份有限公司",
            "short_name": "嘉晶",
            "uid": "16130388",
            "market": "TWSE",
            "raw": {
                "英文簡稱": "EPi",
                "普通股每股面額": "新台幣 10.0000元",
                "已發行普通股數或TDR原股發行股數": "288541819",
            },
        }
        company_query._OFFICIAL_BY_STOCK["3016"] = profile_entry
        company_query._OFFICIAL_BY_UID["16130388"] = profile_entry
        company_query._ISIN_BY_STOCK["3016"] = {
            **profile_entry,
            "isin_code": "TW0003016002",
            "issue_country": "台灣",
            "security_type": "股票",
        }

        with patch(
            "company_query.scrape_company",
            return_value={
                "統一編號": "16130388",
                "登記現況": "核准設立",
                "公司名稱": "嘉晶電子股份有限公司",
                "資本總額(元)": "5000000000",
                "實收資本額(元)": "2885418190",
                "代表人姓名": "徐建華",
                "公司所在地": "新竹科學園區新竹市東區篤行一路10號",
                "登記機關": "國家科學及技術委員會新竹科學園區管理局",
                "核准設立日期": "1998/11/09",
                "最後核准變更日期": "2026/01/20",
                "董監事資料": [],
                "所營事業": "",
                "_share_url": "https://findbiz.nat.gov.tw/fts/query/QueryBar/queryInit.do?banNo=16130388",
                "_error": "findbiz 目前無法直接存取，已改用經濟部商工開放資料",
            },
        ), patch("company_query.get_stock_price_on_or_before", return_value=("2026/05/13", "58.20")), \
           patch("company_query.get_dividends", return_value=[]):
            result = company_query.query_by_uid("16130388", 2026, price_date="2026/05/13")

        self.assertEqual(result["每股金額(元)"], "新台幣 10.0000元")
        self.assertEqual(result["已發行股份總數(股)"], "288541819")
        self.assertEqual(result["章程所訂外文公司名稱"], "EPi")
        self.assertEqual(result["股權狀況"], "開放資料未提供")
        self.assertEqual(
            result["登記資料來源網址"],
            "https://findbiz.nat.gov.tw/fts/query/QueryBar/queryInit.do?banNo=16130388",
        )
        self.assertNotIn("data.gcis.nat.gov.tw/od/data/api", result["登記資料來源網址"])

    def test_search_company_name_falls_back_to_gcis_open_data(self):
        def fake_get(url, **kwargs):
            if "findbiz.nat.gov.tw" in url:
                raise RuntimeError("403 Client Error")

            class _JsonResponse:
                def raise_for_status(self):
                    return None

                def json(self):
                    return [
                        {
                            "Business_Accounting_NO": "03793407",
                            "Company_Name": "臺灣中小企業銀行股份有限公司",
                        }
                    ]

            return _JsonResponse()

        with patch("findbiz_scraper.requests.get", side_effect=fake_get):
            results = findbiz_scraper.search_companies_by_name("臺灣中小企業銀行")

        self.assertEqual(results, [{"ban": "03793407", "name": "臺灣中小企業銀行股份有限公司"}])

    def test_query_by_name_can_resolve_official_short_name(self):
        profile_entry = {
            "stock_no": "2834",
            "name": "臺灣中小企業銀行股份有限公司",
            "short_name": "臺企銀",
            "uid": "03793407",
            "market": "TWSE",
            "raw": {},
        }
        company_query._OFFICIAL_BY_STOCK["2834"] = profile_entry
        company_query._OFFICIAL_BY_UID["03793407"] = profile_entry

        with patch("company_query.query_by_uid", return_value={"統一編號": "03793407", "公司名稱": "臺灣中小企業銀行股份有限公司"}):
            result = company_query.query_by_name("台企銀", 2026, price_date="2026/05/13")

        self.assertEqual(result["統一編號"], "03793407")
        self.assertEqual(result["公司名稱"], "臺灣中小企業銀行股份有限公司")


if __name__ == "__main__":
    unittest.main()
